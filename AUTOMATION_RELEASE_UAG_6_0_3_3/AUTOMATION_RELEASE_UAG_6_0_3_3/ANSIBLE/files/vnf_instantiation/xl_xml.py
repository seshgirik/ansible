from lxml import etree
from openpyxl import Workbook, load_workbook
import xml.dom.minidom
import sys
import re
import cgi

def xmlToExcel(xmlpath):
    encloseWithinCDATA = ['profileXml', 'normalizationXml']
    wb = Workbook()
    worksheet = wb.active
    xmlStr = open(xmlpath).read()
    xmlStr = xmlStr.replace("\t"," ", 1)
    a = xmlStr.find('xmlns="')
    b = xmlStr.find('"', a+len('xmlns="'))
    namespace = xmlStr[a:b+1].replace("xmlns=","").replace('"','')
    xmlStr = xmlStr.replace(" "+xmlStr[a:b+1],"")
    your_tree = etree.fromstring(xmlStr)
    tree = etree.ElementTree(your_tree)
    row = 1
    col = 1
    worksheet.cell(row=row, column=col).value = "namespace"
    worksheet.cell(row=row, column=col+1).value = namespace
    row = row + 1
    for e in your_tree.iter():
        worksheet.cell(row=row, column=col).value = str(tree.getpath(e))
        if str(tree.getpath(e)).split("/")[-1] in encloseWithinCDATA:
            #worksheet.cell(row=row, column=col+1).value = ("<![CDATA[\n"+cgi.escape(e.text.strip())+"\n]]>") if e.text is not None else ''
            worksheet.cell(row=row, column=col+1).value = (e.text.strip().replace("<","&lt;")) if e.text is not None else ''
        else:
            worksheet.cell(row=row, column=col+1).value = e.text.strip() if e.text is not None else ''
        worksheet.cell(row=row, column=col+100).value = str(e.keys())
        worksheet.cell(row=row, column=col+101).value = str(e.values())
        row = row + 1
    savepath = xmlpath.replace(xmlpath.split("/")[-1], "")
    excelname = xmlpath.split("/")[-1].split(".")[0]+".xlsx"
    wb.save(savepath+excelname)


def excelToXml(excelpath):
    wb = load_workbook(filename = excelpath)
    sheet_ranges = wb[wb.sheetnames[0]]
    exitLoop = False
    row_index, col_index = 2, 1
    xml_tree, open_nodes, xml_str = [], [], ""
    namespace = sheet_ranges.cell(1, 2).value

    def get_closed_nodes(xp, open_nodes):
        nodes_present = xp[0].split("/")
        nodes_present.pop(0)
        new_open_nodes, new_close_nodes = [], []
        for node in open_nodes:
            if node in nodes_present:
                new_open_nodes.append(node)
            else:
                a, b = node.find("["), node.find("]")
                if a>0:
                    node = node.replace("["+node[a+1:b]+"]", "")
                new_close_nodes.append(node)
        new_close_nodes.reverse()
        return (new_open_nodes, new_close_nodes)

    while not exitLoop:
        if not sheet_ranges.cell(row_index, col_index).value:
            exitLoop = True
        else:
            xml_tree.append((sheet_ranges.cell(row_index, col_index).value, sheet_ranges.cell(row_index, col_index+1).value))
        row_index = row_index + 1

    for xp in xml_tree:
        nodes_tuple = get_closed_nodes(xp, open_nodes)
        open_nodes = nodes_tuple[0]
        closed_nodes = nodes_tuple[1]
        for node in closed_nodes:
            xml_str = xml_str + "</" + node + ">"
        open_node_name = xp[0].split("/")[len(xp[0].split("/"))-1]
        a, b = open_node_name.find("["), open_node_name.find("]")
        if a>0:
            open_node_name = open_node_name.replace("["+open_node_name[a+1:b]+"]", "")
        if 'NoneType' in str(type(xp[1])):
            xml_str = xml_str + "<" + open_node_name + ">"
            open_nodes.append(xp[0].split("/")[len(xp[0].split("/"))-1])
        else:
            xml_str = xml_str + "<" + open_node_name + ">" + str(xp[1]) + "</" + open_node_name + ">"

    while open_nodes:
        close_node = open_nodes.pop()
        a, b = close_node.find("["), close_node.find("]")
        if a>0:
            close_node = close_node.replace("["+close_node[a+1:b]+"]", "")
        xml_str = xml_str + "</" + close_node + ">"

    xml_str = xml_str.replace("<" + xml_tree[0][0].split("/")[1] + ">", "<" + xml_tree[0][0].split("/")[1] + ' xmlns="' + namespace + '">')
    xml_pr = xml.dom.minidom.parseString(xml_str)
    pretty_xml = xml_pr.toprettyxml()
    pretty_arr = pretty_xml.split("\n")
    pretty_arr.pop(0)
    pretty_xml = "\n".join(pretty_arr)
    savepath = excelpath.replace(excelpath.split("/")[-1], "")
    xmlname = excelpath.split("/")[-1].split(".")[0]+"_New.xml"
    sf = open(savepath+xmlname, "w")
    sf.write(pretty_xml)
    sf.close()

if (__name__=="__main__"):
    try:
        path = sys.argv[1]
        if len(sys.argv)==3:
            xmlStr = open(path).read()
            xmlStr = xmlStr.replace("\t"," ", 1)
            a = xmlStr.find('xmlns="')
            b = xmlStr.find('"', a+len('xmlns="'))
            namespace = xmlStr[a:b+1].replace("xmlns=","").replace('"','')
            xmlStr = xmlStr.replace(" "+xmlStr[a:b+1],"")
            your_tree = etree.fromstring(xmlStr)
            tree = etree.ElementTree(your_tree)
            for e in your_tree.iter():
                if "vdu-depends-on" in str(tree.getpath(e)):
                    print tree.getpath(e)
        else:
            if path.split(".")[-1]=="xml":
                xmlToExcel(path)
            elif path.split(".")[-1]=="xlsx":
                excelToXml(path)
            else:
                print ("ERROR: Unsupported file")
    except:
        print "\n\nUSAGE: python xl_xml.py /path/to/excel_or_xml_file\n\nMake sure the file is present in path specified\n\n\n"
